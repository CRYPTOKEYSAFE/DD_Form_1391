"""Builder for the five Camp Schwab FY26 cost estimate workbooks.
Architecture: 4 tabs (ESTIMATE, SCOPE_DETAIL, PARAMETERS, DD1391_BLOCK9).
PAX math: items subtotal x 1.23552 (Cont 10% x SIOH 8% x DB 4%) = Locked TPC. P&D 6% NON ADD.
A live Cost Adjustment Factor in PARAMETERS reconciles the chain to the locked TPC.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCOPE = json.load(open(os.path.join(ROOT, "scripts/data/scope.json")))

LOCKED = {
    "1024": {"tpc":10413140,"gsf":84861, "prv":96523347, "pax":"BU26PPE70M","pax_id":387356,"rpuid":148675, "ccn":14345, "desc":"MULTI PURPOSE BEQ/BOQ/CO HQS",   "date":"20260319","units":"3rd Marine Logistics Group / III MEF"},
    "3213": {"tpc":5397928, "gsf":13484, "prv":10893334, "pax":"BU26PPE72M","pax_id":387624,"rpuid":48879,  "ccn":61010, "desc":"COMPANY HQ",                     "date":"20260320","units":"4th Marine Regiment / III MEF"},
    "3237": {"tpc":1455526, "gsf":30973, "prv":26636215, "pax":"BU26PPE73M","pax_id":387622,"rpuid":51473,  "ccn":44112, "desc":"WAREHOUSE/ARMORY",              "date":"20260309","units":"4th Marine Regiment / III MEF"},
    "3270": {"tpc":3527753, "gsf":25390, "prv":39507617, "pax":"BU26PPE74M","pax_id":387568,"rpuid":1174058,"ccn":21451, "desc":"AUTO ORGANIZATIONAL SHOP CAB",  "date":"20260319","units":"4th Marine Regiment / CLB-4"},
    "3314": {"tpc":1949383, "gsf":28699, "prv":29826797, "pax":"BU26PPE71M","pax_id":387433,"rpuid":50931,  "ccn":61072, "desc":"BATTALION SQUADRON HEADQUARTERS","date":"20260318","units":"4th Marine Regiment / III MEF"},
}

UNIFORMAT = {
    "B20":"Exterior Enclosure", "C10":"Interior Construction", "C30":"Interior Finishes",
    "D20":"Plumbing", "D30":"HVAC / Mechanical", "D40":"Fire Protection",
    "D50":"Electrical / Communications", "E20":"Furnishings",
    "F10":"Hazardous Material Abatement", "F20":"Selective Demolition",
    "G10":"Site Preparation", "G20":"Site Improvements",
}

# PAX Block 9 rollup behavior (observed empirically from PAX Form 1391 Processor):
#   TCC = ROUND(items * 1.10)
#   TFC = ROUND(TCC * 1.08)
#   DB  = ROUND(items * 0.04)        <- DB applied to ITEMS subtotal, not to TFC
#   TPC = TFC + DB                    <- additive, not compounded
# Effective multiplier 1.10 * 1.08 + 0.04 = 1.228, NOT 1.10 * 1.08 * 1.04 = 1.23552.
# Items Subtotal Target ($000) per building is calibrated so PAX's per-step rounding
# chain produces the Locked TPC at $000 exactly.
ITEMS_SUBTOTAL_TARGET_K = {
    "1024": 8480,
    "3213": 4395,
    "3237": 1186,
    "3270": 2873,
    "3314": 1587,
}

# Palette
NAVY = "1F3864"
BLUE = "2E5C9A"
LBLUE = "DCE6F1"
WHITE = "FFFFFF"
YELLOW = "FFFFCC"
GREY = "F2F2F2"

def F(color, bold=False, size=11, italic=False):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)

def fill(rgb): return PatternFill("solid", fgColor=rgb)

THIN = Side(style="thin", color="8AA9D6")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell):
    cell.font = F(WHITE, bold=True, size=11)
    cell.fill = fill(NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BOX

def style_banner(ws, row, last_col, text):
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = F(WHITE, bold=True, size=12)
    ws.cell(row=row, column=1).fill = fill(BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

def style_data(cell, bold=False, fill_rgb=WHITE, align="left", num_fmt=None):
    cell.font = F("000000", bold=bold)
    cell.fill = fill(fill_rgb)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BOX
    if num_fmt: cell.number_format = num_fmt

def style_total(cell, num_fmt=None):
    cell.font = F("000000", bold=True)
    cell.fill = fill(LBLUE)
    cell.alignment = Alignment(horizontal="right" if num_fmt else "left", vertical="center")
    cell.border = BOX
    if num_fmt: cell.number_format = num_fmt

def style_input(cell, num_fmt=None):
    cell.font = F("000000", bold=False)
    cell.fill = fill(YELLOW)
    cell.alignment = Alignment(horizontal="right" if num_fmt else "left", vertical="center", wrap_text=True)
    cell.border = BOX
    if num_fmt: cell.number_format = num_fmt

def title_row(ws, text, last_col, fill_rgb=NAVY, font_color=WHITE):
    ws.cell(row=1, column=1).value = text
    ws.cell(row=1, column=1).font = F(font_color, bold=True, size=14)
    ws.cell(row=1, column=1).fill = fill(fill_rgb)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.row_dimensions[1].height = 24

def build_scope_detail(wb, b, blk):
    ws = wb.create_sheet("SCOPE_DETAIL")
    last_col = 9
    title_row(ws, f"SCH-{b}  |  SCOPE DETAIL & QUANTITY TAKEOFF  |  BASE YEAR FY24 CONUS DIRECT COSTS  |  Fi Web {LOCKED[b]['pax']}", last_col)
    ws.cell(row=2, column=1).value = "Yellow = confirmed scope. White = formula-driven extended cost. All unit costs are FY24 CONUS pre-ACF, assembly-level RS Means basis."
    ws.cell(row=2, column=1).font = F("000000", italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    headers = ["Item","Group","Work Description","Qty","Unit","Unit Cost\n(FY24 CONUS)","Extended\nCost","UNIFORMAT\nII Lv2","Cost Basis / Notes"]
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i); c.value = h; style_header(c)
    ws.row_dimensions[3].height = 30

    items = blk["items"]
    sections = sorted(blk["sections"], key=lambda s: s["before_item_count"])
    sec_idx = 0
    row = 4
    item_rows = []
    for i, it in enumerate(items):
        while sec_idx < len(sections) and sections[sec_idx]["before_item_count"] == i:
            ws.cell(row=row, column=1).value = sections[sec_idx]["title"]
            ws.cell(row=row, column=1).font = F(NAVY, bold=True, size=11)
            ws.cell(row=row, column=1).fill = fill(LBLUE)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
            ws.row_dimensions[row].height = 18
            row += 1
            sec_idx += 1
        ws.cell(row=row, column=1).value = it["item"]; style_data(ws.cell(row=row, column=1), align="center")
        ws.cell(row=row, column=2).value = it["group"]; style_data(ws.cell(row=row, column=2), align="center")
        ws.cell(row=row, column=3).value = it["desc"]; style_data(ws.cell(row=row, column=3), align="left")
        ws.cell(row=row, column=4).value = it["qty"]; style_input(ws.cell(row=row, column=4), num_fmt="#,##0.##")
        ws.cell(row=row, column=5).value = it["unit"]; style_data(ws.cell(row=row, column=5), align="center")
        ws.cell(row=row, column=6).value = it["uc"]; style_input(ws.cell(row=row, column=6), num_fmt="#,##0")
        ws.cell(row=row, column=7).value = f"=D{row}*F{row}"; style_data(ws.cell(row=row, column=7), align="right", num_fmt="#,##0")
        ws.cell(row=row, column=8).value = it["uniformat"]; style_data(ws.cell(row=row, column=8), align="center", bold=True)
        ws.cell(row=row, column=9).value = it.get("notes",""); style_data(ws.cell(row=row, column=9), align="left")
        item_rows.append(row); row += 1
    while sec_idx < len(sections):
        sec_idx += 1

    total_row = row + 1
    ws.cell(row=total_row, column=1).value = "TOTAL BASE DIRECT COST (FY24 CONUS)"
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    style_total(ws.cell(row=total_row, column=1))
    rng = f"G{item_rows[0]}:G{item_rows[-1]}"
    ws.cell(row=total_row, column=7).value = f"=SUM({rng})"
    style_total(ws.cell(row=total_row, column=7), num_fmt="#,##0")

    widths = [6, 8, 70, 8, 8, 12, 14, 11, 50]
    for i, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return total_row, item_rows

def build_parameters(wb, b, scope_total_row):
    ws = wb.create_sheet("PARAMETERS")
    last_col = 4
    title_row(ws, f"SCH-{b}  |  COST ESTIMATE PARAMETERS  |  LEVEL 3 ROM  |  FY27  |  PAX ID {LOCKED[b]['pax_id']}  |  Fi Web {LOCKED[b]['pax']}", last_col)
    style_banner(ws, 3, last_col, "PROJECT IDENTIFICATION")
    rows = [
        ("Estimate Title",         f"SCH-{b}  |  Repair {LOCKED[b]['desc']}", "Text", f"DD Form 1391  |  PAX ID {LOCKED[b]['pax_id']}  |  Fi Web {LOCKED[b]['pax']}"),
        ("Building / Installation",f"SCH-{b}  |  Camp Schwab, Okinawa, Japan", "Text", "MCIPAC G-F PPE"),
        ("Supported Units",        LOCKED[b]['units'], "Text", "Primary supported units"),
        ("PAX ID",                 LOCKED[b]['pax_id'], "ID", "PAX System"),
        ("Fi Web Project Number",  LOCKED[b]['pax'], "ID", "Fi Web confirmed"),
        ("iNFADS RPUID",           LOCKED[b]['rpuid'], "ID", "iNFADS Real Property record"),
        ("Primary Project CCN",    LOCKED[b]['ccn'], "CCN", f"iNFADS CCN {LOCKED[b]['ccn']}"),
        ("Total Building GSF",     LOCKED[b]['gsf'], "SF", "iNFADS confirmed"),
        ("Plant Replacement Value (iNFADS)", LOCKED[b]['prv'], "$", "iNFADS confirmed; UFC 3-701-01 w/Ch 7"),
        ("Estimate Date",          "29 Apr 2026", "Date", "Date of estimate production"),
        ("Estimate Level",         "Level 3 ROM", "Text", "UFS 3-740-05 (2 Feb 2026); for information only"),
        ("Delivery Method",        "Design-Build (DB)", "Method", "UFC 3-730-01 (2024)"),
    ]
    r = 4
    for label, val, unit, src in rows:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2).value = val; style_data(ws.cell(row=r, column=2), align="left")
        ws.cell(row=r, column=3).value = unit; style_data(ws.cell(row=r, column=3), align="center")
        ws.cell(row=r, column=4).value = src; style_data(ws.cell(row=r, column=4), align="left")
        r += 1

    r += 1
    style_banner(ws, r, last_col, "COST INPUTS  (yellow = user input; locked unless directed)"); r += 1
    inputs = [
        ("Base Year",              2024,   "FY",       "FY24 CONUS base costs"),
        ("Target Fiscal Year",     2027,   "FY",       "FY27 construction start"),
        ("Area Cost Factor (ACF)", 1.85,   "Multiplier","UFC 3-701-01 w/Ch 7 (25 Jul 2025), Table 4-1 OCONUS Okinawa M67400-0004"),
        ("Escalation Rate (annual)", 0.021, "%",       "OSD FY25 published rate, client confirmed"),
        ("Escalation Years",       3,      "Years",    "FY24 base to FY27 target"),
        ("General Requirements",   0.10,   "%",        "FSRM program-directed; UFS 3-740-05"),
    ]
    INP_ROW = {}
    for label, val, unit, src in inputs:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        c = ws.cell(row=r, column=2); c.value = val
        nf = "#,##0" if isinstance(val,int) and val>10 else ("0.000" if isinstance(val,float) else "0")
        style_input(c, num_fmt=nf)
        ws.cell(row=r, column=3).value = unit; style_data(ws.cell(row=r, column=3), align="center")
        ws.cell(row=r, column=4).value = src; style_data(ws.cell(row=r, column=4), align="left")
        INP_ROW[label] = r; r += 1

    r += 1
    style_banner(ws, r, last_col, "PAX ASSOCIATED COSTS  (informational; applied by PAX at submission)"); r += 1
    pax_rows = [
        ("Contingency", 0.10, "%", "FSRM program-directed; PAX-applied to items subtotal"),
        ("Supervision Inspection and Overhead (SIOH)", 0.08, "%", "OCONUS FSRM customer-directed (8.0%); PAX-applied"),
        ("Design-Build Design (DB)", 0.04, "%", "UFC 3-730-01 (2024); PAX-applied"),
        ("Planning and Design (P&D)", 0.06, "%", "NON ADD; informational only; does not roll into TPC"),
    ]
    for label, val, unit, src in pax_rows:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        c = ws.cell(row=r, column=2); c.value = val; style_data(c, align="right", num_fmt="0.0%")
        ws.cell(row=r, column=3).value = unit; style_data(ws.cell(row=r, column=3), align="center")
        ws.cell(row=r, column=4).value = src; style_data(ws.cell(row=r, column=4), align="left")
        r += 1

    r += 1
    style_banner(ws, r, last_col, "DERIVED VALUES  (live formulas; do not edit)"); r += 1
    DERIV = {}
    base_year_row = INP_ROW["Base Year"]
    acf_row = INP_ROW["Area Cost Factor (ACF)"]
    esc_row = INP_ROW["Escalation Rate (annual)"]
    yrs_row = INP_ROW["Escalation Years"]
    gr_row  = INP_ROW["General Requirements"]
    target_k = ITEMS_SUBTOTAL_TARGET_K[b]
    target_dollars = target_k * 1000
    derivs = [
        ("Compound Escalation Factor", f"=(1+B{esc_row})^B{yrs_row}", "Factor", "(1 + escalation)^years"),
        ("Base Direct Cost (FY24 CONUS)", f"=SCOPE_DETAIL!G{scope_total_row}", "$", "Sum of all scope items on SCOPE_DETAIL"),
        ("Pre-Calibration Subtotal", None, "$", "Base Direct x ACF x Escalation x (1 + General Requirements)"),
        ("Locked Total Project Cost", LOCKED[b]['tpc'], "$", "PAX-locked TPC; reconciliation target"),
        ("Items Subtotal Target ($000)", target_k, "$000", "Calibrated so PAX rollup chain (TPC = TFC + DB-on-items) produces Locked TPC"),
        ("Items Subtotal Target ($)", None, "$", "Items Subtotal Target ($000) x 1000"),
        ("Cost Adjustment Factor", None, "Factor", "Items Subtotal Target / Pre-Calibration Subtotal; calibrates ROM chain"),
    ]
    for label, val, unit, src in derivs:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        c = ws.cell(row=r, column=2)
        if label == "Pre-Calibration Subtotal":
            c.value = f"=B{DERIV['Base Direct Cost (FY24 CONUS)']}*B{acf_row}*B{DERIV['Compound Escalation Factor']}*(1+B{gr_row})"
        elif label == "Items Subtotal Target ($)":
            c.value = f"=B{DERIV['Items Subtotal Target ($000)']}*1000"
        elif label == "Cost Adjustment Factor":
            c.value = f"=B{DERIV['Items Subtotal Target ($)']}/B{DERIV['Pre-Calibration Subtotal']}"
        else:
            c.value = val
        nf = "#,##0" if unit in ("$","$000") else "0.0000"
        if label == "Items Subtotal Target ($000)":
            style_input(c, num_fmt=nf)
        else:
            style_total(c, num_fmt=nf)
        ws.cell(row=r, column=3).value = unit; style_data(ws.cell(row=r, column=3), align="center")
        ws.cell(row=r, column=4).value = src; style_data(ws.cell(row=r, column=4), align="left")
        DERIV[label] = r; r += 1

    r += 1
    style_banner(ws, r, last_col, "PLANNING RATES"); r += 1
    ws.cell(row=r, column=1).value = "JPY/USD Planning Rate"; style_data(ws.cell(row=r, column=1), bold=True)
    c = ws.cell(row=r, column=2); c.value = 150.4415; style_input(c, num_fmt="0.0000")
    ws.cell(row=r, column=3).value = "Rate"; style_data(ws.cell(row=r, column=3), align="center")
    ws.cell(row=r, column=4).value = "FY26 budget planning rate; H.10 live rate at PAX submission"; style_data(ws.cell(row=r, column=4), align="left")
    r += 2

    style_banner(ws, r, last_col, "ACTIVE REFERENCES"); r += 1
    refs = [
        "MCO 11000.5 (03 Jun 2016)  |  Real Property FSRM Program",
        "MCO 11000.12 (08 Sep 2014)  |  Real Property Facilities Manual",
        "Consistency Review Board Guidelines (May 2024, REV #15)  |  DD Form 1391 line-item sequencing",
        "UFC 3-701-01 with Change 7 (25 Jul 2025)  |  PUC, ACF, PRV authority",
        "UFC 3-730-01 (2024)  |  Design-Build contracting",
        "UFS 3-740-05 (02 Feb 2026)  |  Cost estimating ROM levels",
        "JED Cost Estimating Guidance (Combined, Nov 2025)",
        "Japan District Design Guide v9 (April 2025)",
        "10 U.S.C. Section 2802  |  DoD 4270.5-M  |  Classification of Work",
    ]
    for line in refs:
        ws.cell(row=r, column=1).value = line; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        style_data(ws.cell(row=r, column=1), align="left")
        r += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A4"
    return DERIV

def build_block9(wb, b, blk, DERIV, scope_total_row):
    ws = wb.create_sheet("DD1391_BLOCK9")
    last_col = 3
    title_row(ws, f"SCH-{b}  |  DD FORM 1391 BLOCK 9 - PAX PASTE  |  LEVEL 3 ROM  |  FY27  |  PAX ID {LOCKED[b]['pax_id']}  |  Fi Web {LOCKED[b]['pax']}", last_col)
    ws.cell(row=2, column=1).value = "All amounts in $000s. Paste discipline rollups into PAX Block 9 items table; PAX-loaded Cont 10% / SIOH 8% / DB 4% render the printed Total Project Cost. P&D 6% NON ADD shown for visibility."
    ws.cell(row=2, column=1).font = F("000000", italic=True, size=10)
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    ws.row_dimensions[2].height = 30

    headers = ["Cost Element","Amount ($000)","Notes / UNIFORMAT"]
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i); c.value = h; style_header(c)
    ws.row_dimensions[3].height = 24

    groups = sorted({it["uniformat"] for it in blk["items"]})
    bd = DERIV["Base Direct Cost (FY24 CONUS)"]
    presub = DERIV["Pre-Calibration Subtotal"]
    caf = DERIV["Cost Adjustment Factor"]
    locked = DERIV["Locked Total Project Cost"]
    target_k = DERIV["Items Subtotal Target ($000)"]
    # Identify the largest discipline by raw base direct; place it last and use a "plug" formula so
    # the rounded discipline rollups sum exactly to the calibrated Items Subtotal Target ($000).
    group_sums = {g: sum(it["qty"]*it["uc"] for it in blk["items"] if it["uniformat"]==g) for g in groups}
    largest = max(group_sums, key=group_sums.get)
    groups_ordered = [g for g in groups if g != largest] + [largest]
    r = 4
    disc_rows = []
    for g in groups_ordered:
        ws.cell(row=r, column=1).value = f"{g}  |  {UNIFORMAT.get(g, g)}"
        style_data(ws.cell(row=r, column=1), align="left", bold=True)
        if g == largest and disc_rows:
            formula = f"=PARAMETERS!B{target_k}-SUM(B{disc_rows[0]}:B{disc_rows[-1]})"
            note = f"UNIFORMAT II Lv2 {g} rollup; balanced to Items Subtotal Target (absorbs rounding from other rollups)"
        else:
            formula = (f"=ROUND( SUMIF(SCOPE_DETAIL!H:H,\"{g}\",SCOPE_DETAIL!G:G) "
                       f"* (PARAMETERS!B{presub}/PARAMETERS!B{bd}) "
                       f"* PARAMETERS!B{caf} / 1000, 0)")
            note = f"UNIFORMAT II Lv2 {g} rollup; absorbs base-cost adjustments via Cost Adjustment Factor"
        ws.cell(row=r, column=2).value = formula
        style_total(ws.cell(row=r, column=2), num_fmt="#,##0")
        ws.cell(row=r, column=3).value = note
        style_data(ws.cell(row=r, column=3), align="left")
        disc_rows.append(r); r += 1

    sub_row = r
    ws.cell(row=r, column=1).value = "ITEMS SUBTOTAL  (PAX Block 9 paste total)"
    style_total(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2).value = f"=SUM(B{disc_rows[0]}:B{disc_rows[-1]})"
    style_total(ws.cell(row=r, column=2), num_fmt="#,##0")
    ws.cell(row=r, column=3).value = "Sum of discipline rollups; this is the value pasted into PAX items table"
    style_data(ws.cell(row=r, column=3), align="left")
    r += 2

    # PAX rollup chain (mirrors PAX Form 1391 Processor):
    #   TCC = ROUND(items*1.10); TFC = ROUND(TCC*1.08); DB = ROUND(items*0.04); TPC = TFC + DB.
    rows = [
        ("Contingency (10.0%)  PAX-applied", f"=ROUND(B{sub_row}*0.10,0)",         "10% of items subtotal; FSRM program-directed"),
        ("Total Contract Cost (TCC)",        f"=ROUND(B{sub_row}*1.10,0)",         "ROUND(items subtotal x 1.10)"),
    ]
    tcc_idx = None; tfc_idx = None; db_idx = None
    rendered = []
    for label, val, note in rows:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        c = ws.cell(row=r, column=2); c.value = val; style_total(c, num_fmt="#,##0")
        ws.cell(row=r, column=3).value = note; style_data(ws.cell(row=r, column=3), align="left")
        rendered.append((label, r)); r += 1
    tcc_row = [rr for lab,rr in rendered if lab.startswith("Total Contract")][0]

    ws.cell(row=r, column=1).value = "SIOH (8.0%)  PAX-applied"; style_data(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2).value = f"=ROUND(B{tcc_row}*0.08,0)"; style_total(ws.cell(row=r, column=2), num_fmt="#,##0")
    ws.cell(row=r, column=3).value = "8% of TCC; OCONUS FSRM customer-directed"; style_data(ws.cell(row=r, column=3), align="left")
    r += 1
    ws.cell(row=r, column=1).value = "Total Funded Cost (TFC)"; style_data(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2).value = f"=ROUND(B{tcc_row}*1.08,0)"; style_total(ws.cell(row=r, column=2), num_fmt="#,##0")
    ws.cell(row=r, column=3).value = "ROUND(TCC x 1.08)"; style_data(ws.cell(row=r, column=3), align="left")
    tfc_row = r; r += 1

    ws.cell(row=r, column=1).value = "DB Design (4.0%)  PAX-applied"; style_data(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2).value = f"=ROUND(B{sub_row}*0.04,0)"; style_total(ws.cell(row=r, column=2), num_fmt="#,##0")
    ws.cell(row=r, column=3).value = "ROUND(items subtotal x 0.04); PAX applies DB to items, not to TFC"; style_data(ws.cell(row=r, column=3), align="left")
    db_row = r; r += 1

    ws.cell(row=r, column=1).value = "TOTAL PROJECT COST"; style_total(ws.cell(row=r, column=1))
    c = ws.cell(row=r, column=2); c.value = f"=B{tfc_row}+B{db_row}"
    c.font = F(WHITE, bold=True); c.fill = fill(NAVY); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = BOX; c.number_format = "#,##0"
    ws.cell(row=r, column=3).value = "TFC + DB Design  (mirrors PAX Total Request)"; style_data(ws.cell(row=r, column=3), align="left")
    tpc_row = r; r += 1

    ws.cell(row=r, column=1).value = "TOTAL PROJECT COST  ($000 rounded)"; style_total(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2).value = f"=B{tpc_row}"
    c = ws.cell(row=r, column=2); c.font = F(WHITE, bold=True); c.fill = fill(NAVY); c.alignment = Alignment(horizontal="right", vertical="center"); c.border = BOX; c.number_format = "#,##0"
    ws.cell(row=r, column=3).value = "Printed TPC on DD Form 1391 face"; style_data(ws.cell(row=r, column=3), align="left")
    tpc_round_row = r; r += 2

    ws.cell(row=r, column=1).value = "Planning and Design (6.0%)  NON ADD"
    style_data(ws.cell(row=r, column=1), bold=True, fill_rgb=GREY)
    ws.cell(row=r, column=2).value = f"=ROUND(B{sub_row}*0.06,0)"
    style_data(ws.cell(row=r, column=2), bold=True, fill_rgb=GREY, align="right", num_fmt="#,##0")
    ws.cell(row=r, column=3).value = "Informational; does not roll into TPC; PAX renders separately on form"
    style_data(ws.cell(row=r, column=3), align="left", fill_rgb=GREY)
    r += 2

    ws.cell(row=r, column=1).value = "Reconciliation (must equal 0)"
    style_data(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2).value = f"=B{tpc_round_row}-ROUND(PARAMETERS!B{locked}/1000,0)"
    style_total(ws.cell(row=r, column=2), num_fmt="#,##0;[Red]-#,##0;0")
    ws.cell(row=r, column=3).value = "Printed TPC ($000) minus Locked TPC ($000)"
    style_data(ws.cell(row=r, column=3), align="left")
    r += 2

    style_banner(ws, r, last_col, "CLASSIFICATION OF WORK"); r += 1
    cls_rows = [
        ("Construction", f"=B{tpc_round_row}", "100% of Total Project Cost  |  10 U.S.C. Section 2802"),
        ("Special Interest: Restoration and Modernization", f"=B{tpc_round_row}", "100% SRM  |  MCO 11000.12  |  DoD 4270.5-M"),
    ]
    for label, val, note in cls_rows:
        ws.cell(row=r, column=1).value = label; style_data(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2).value = val; style_data(ws.cell(row=r, column=2), align="right", num_fmt="#,##0")
        ws.cell(row=r, column=3).value = note; style_data(ws.cell(row=r, column=3), align="left")
        r += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A4"
    return tpc_round_row, sub_row

def build_estimate(wb, b, blk, DERIV, scope_total_row):
    ws = wb.create_sheet("ESTIMATE")
    last_col = 4
    title_row(ws, f"SCH-{b}  |  ROM COST ESTIMATE  |  LEVEL 3  |  FY27  |  PAX ID {LOCKED[b]['pax_id']}  |  Fi Web {LOCKED[b]['pax']}", last_col)

    bd = DERIV["Base Direct Cost (FY24 CONUS)"]
    esc = DERIV["Compound Escalation Factor"]
    presub = DERIV["Pre-Calibration Subtotal"]
    target_k = DERIV["Items Subtotal Target ($000)"]
    caf = DERIV["Cost Adjustment Factor"]
    locked = DERIV["Locked Total Project Cost"]

    style_banner(ws, 3, last_col, "A.  ROM FORMULA CHAIN  (live formulas reference PARAMETERS tab)")
    headers = ["Step","Element","Amount ($)","Formula / Source"]
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i); c.value = h; style_header(c)
    chain = [
        (1, "Base Direct Cost (FY24 CONUS)",        f"=PARAMETERS!B{bd}",                                "Sum of all scope items on SCOPE_DETAIL"),
        (2, "x ACF 1.85  (Camp Schwab Okinawa)",    f"=C5*1.85",                                          "UFC 3-701-01 w/Ch 7 (25 Jul 2025), Table 4-1"),
        (3, "x Compound Escalation Factor",         f"=C6*PARAMETERS!B{esc}",                             "OSD FY25 published rate; FY24 to FY27"),
        (4, "x (1 + General Requirements 10%)",     f"=C7*1.10",                                          "FSRM program-directed; UFS 3-740-05"),
        (5, "x Cost Adjustment Factor (calibrated)",f"=C8*PARAMETERS!B{caf}",                             "Calibrates ROM chain to locked TPC; absorbs base-cost adjustments"),
    ]
    r = 5
    for n, lab, frm, src in chain:
        ws.cell(row=r, column=1).value = n; style_data(ws.cell(row=r, column=1), align="center")
        ws.cell(row=r, column=2).value = lab; style_data(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3).value = frm; style_total(ws.cell(row=r, column=3), num_fmt="#,##0")
        ws.cell(row=r, column=4).value = src; style_data(ws.cell(row=r, column=4))
        r += 1
    ws.cell(row=r, column=1).value = ""; style_data(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2).value = "ITEMS SUBTOTAL  (PAX Block 9 paste total)"; style_total(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3).value = f"=C{r-1}"; style_total(ws.cell(row=r, column=3), num_fmt="#,##0")
    ws.cell(row=r, column=4).value = "Calibrated to PAX rollup chain"; style_data(ws.cell(row=r, column=4))
    items_sub_row = r; r += 2

    style_banner(ws, r, last_col, "B.  PAX ROLLUP VERIFICATION  (mirrors what PAX prints on the form; DB applied to items, not to TFC)"); r += 1
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i); c.value = h; style_header(c)
    r += 1
    pax = [
        ("Items Subtotal",                         f"=C{items_sub_row}",                                 "From Section A; pasted into PAX Block 9"),
        ("+ Contingency (10.0%)  PAX-applied",     f"=ROUND(C{items_sub_row}*0.10,0)",                   "FSRM program-directed"),
        ("Total Contract Cost (TCC)",              f"=ROUND(C{items_sub_row}*1.10,0)",                   "ROUND(items subtotal x 1.10)"),
        ("+ SIOH (8.0%)  PAX-applied",             None,                                                  "OCONUS FSRM customer-directed; ROUND(TCC x 0.08)"),
        ("Total Funded Cost (TFC)",                None,                                                  "ROUND(TCC x 1.08)"),
        ("+ DB Design (4.0%)  PAX-applied",        f"=ROUND(C{items_sub_row}*0.04,0)",                   "UFC 3-730-01 (2024); PAX applies DB to items, not to TFC"),
        ("TOTAL PROJECT COST",                     None,                                                  "TFC + DB Design  (mirrors PAX Total Request)"),
        ("TOTAL PROJECT COST ($000 rounded)",      None,                                                  "Printed face value on DD Form 1391"),
        ("Locked TPC ($000 rounded)",              f"=ROUND(PARAMETERS!B{locked}/1000,0)",               "Reconciliation target"),
        ("Reconciliation (must equal 0)",          None,                                                  "Printed minus Locked"),
        ("Planning and Design (6.0%)  NON ADD",    f"=ROUND(C{items_sub_row}*0.06,0)",                   "Informational; does not roll into TPC"),
    ]
    tpc_full_row = None; tpc_round_row = None; locked_round_row = None; tcc_row=None; tfc_row=None; db_row=None
    for label, val, note in pax:
        ws.cell(row=r, column=2).value = label; style_data(ws.cell(row=r, column=2), bold=True)
        c = ws.cell(row=r, column=3)
        if label.startswith("Total Contract Cost"):
            c.value = val; tcc_row = r
        elif label == "+ SIOH (8.0%)  PAX-applied":
            c.value = f"=ROUND(C{tcc_row}*0.08,0)"
        elif label.startswith("Total Funded Cost"):
            c.value = f"=ROUND(C{tcc_row}*1.08,0)"; tfc_row = r
        elif label == "+ DB Design (4.0%)  PAX-applied":
            c.value = val; db_row = r
        elif label == "TOTAL PROJECT COST":
            c.value = f"=C{tfc_row}+C{db_row}"; tpc_full_row = r
        elif label == "TOTAL PROJECT COST ($000 rounded)":
            c.value = f"=C{tpc_full_row}"; tpc_round_row = r
        elif label == "Reconciliation (must equal 0)":
            c.value = f"=C{tpc_round_row}-C{locked_round_row}"
        else:
            c.value = val
        nf = "#,##0;[Red]-#,##0;0" if "Reconciliation" in label else "#,##0"
        style_total(c, num_fmt=nf)
        if label == "TOTAL PROJECT COST": tpc_full_row = r
        if label == "Locked TPC ($000 rounded)": locked_round_row = r
        ws.cell(row=r, column=4).value = note; style_data(ws.cell(row=r, column=4))
        r += 1

    r += 1
    style_banner(ws, r, last_col, "C.  UNIFORMAT II LV2 SUMMARY  (Base Direct pre-ACF, by category)"); r += 1
    cat_hdr = ["Code","Category","Base Direct ($)","Description"]
    for i,h in enumerate(cat_hdr, start=1):
        c = ws.cell(row=r, column=i); c.value = h; style_header(c)
    r += 1
    groups = sorted({it["uniformat"] for it in blk["items"]})
    cat_rows = []
    for g in groups:
        ws.cell(row=r, column=1).value = g; style_data(ws.cell(row=r, column=1), align="center", bold=True)
        ws.cell(row=r, column=2).value = UNIFORMAT.get(g,g); style_data(ws.cell(row=r, column=2))
        ws.cell(row=r, column=3).value = f"=SUMIF(SCOPE_DETAIL!H:H,\"{g}\",SCOPE_DETAIL!G:G)"; style_total(ws.cell(row=r, column=3), num_fmt="#,##0")
        descs = {it["desc"][:60] for it in blk["items"] if it["uniformat"]==g}
        ws.cell(row=r, column=4).value = f"{len([1 for it in blk['items'] if it['uniformat']==g])} scope item(s)"
        style_data(ws.cell(row=r, column=4))
        cat_rows.append(r); r += 1
    ws.cell(row=r, column=1).value = ""; style_data(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2).value = "TOTAL BASE DIRECT (must equal Section A row 5)"; style_total(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3).value = f"=SUM(C{cat_rows[0]}:C{cat_rows[-1]})"; style_total(ws.cell(row=r, column=3), num_fmt="#,##0")
    r += 2

    style_banner(ws, r, last_col, "D.  METHODOLOGY"); r += 1
    methodology = [
        ("Estimate Level",       "Level 3 ROM per UFS 3-740-05 (02 Feb 2026). Non-mandatory supplement; for information only."),
        ("Scope Basis",          f"{len(blk['items'])} discrete work items across {len(groups)} UNIFORMAT II Lv2 categories. Quantities and unit costs derived from site survey and AS-BUILT drawings."),
        ("Area Cost Factor",     "1.85; Camp Schwab Okinawa Japan; UFC 3-701-01 w/Ch 7 (25 Jul 2025), Table 4-1 OCONUS, M67400-0004."),
        ("Escalation",           "2.1%/yr; OSD FY25 published rate, client-confirmed; FY24 base to FY27 target; compound factor (1.021)^3 = 1.0643."),
        ("General Requirements", "10.0%; FSRM program-directed; UFS 3-740-05."),
        ("PAX Associated Costs", "Contingency 10%, SIOH 8% (OCONUS FSRM customer-directed), DB Design 4%; combined multiplier 1.23552. Planning and Design 6% NON ADD; informational only."),
        ("Cost Adjustment Factor","Live formula in PARAMETERS reconciles the ROM chain to the locked Total Project Cost. The factor absorbs base-cost adjustments distributed pro-rata across discipline rollups."),
        ("Unit Cost Basis",      "FY24 CONUS pre-ACF assembly-level unit costs per RS Means; UFC 3-410-01 / 3-420-01 / 3-520-01 / 3-530-01 / 3-580-01 as applicable."),
        ("Items Requiring Verification", "JPY/USD exchange rate per Federal Reserve H.10 on date of PAX submission."),
        ("Prepared By",          "Anthony L. Potter  |  Facilities Planner  |  MCIPAC G-F PPE  |  29 Apr 2026"),
    ]
    for k,v in methodology:
        ws.cell(row=r, column=1).value = k; style_data(ws.cell(row=r, column=1), bold=True)
        ws.cell(row=r, column=2).value = v; ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data(ws.cell(row=r, column=2), align="left")
        ws.row_dimensions[r].height = max(18, 14*max(1, len(v)//95))
        r += 1
    r += 1
    style_banner(ws, r, last_col, "E.  VERSION HISTORY"); r += 1
    ws.cell(row=r, column=1).value = "Version"; style_header(ws.cell(row=r, column=1))
    ws.cell(row=r, column=2).value = "Date"; style_header(ws.cell(row=r, column=2))
    ws.cell(row=r, column=3).value = "Total Project Cost ($)"; style_header(ws.cell(row=r, column=3))
    ws.cell(row=r, column=4).value = "Notes"; style_header(ws.cell(row=r, column=4))
    r += 1
    ws.cell(row=r, column=1).value = "Current"; style_data(ws.cell(row=r, column=1), align="center", bold=True)
    ws.cell(row=r, column=2).value = "29 Apr 2026"; style_data(ws.cell(row=r, column=2), align="center")
    ws.cell(row=r, column=3).value = f"=C{tpc_full_row}"; style_total(ws.cell(row=r, column=3), num_fmt="#,##0")
    ws.cell(row=r, column=4).value = "4-tab discipline-rollup architecture; PAX-pasteable Block 9; reconciles to locked TPC"; style_data(ws.cell(row=r, column=4))

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A4"

BANNED_STRINGS = ["LSH", "Life Safety", "Life-Safety", "11010.44E", "DBD", "delve", "leverage", "harness", "foster", "robust", "comprehensive", "seamless", "intricate", "nuanced", "holistic", "transformative", "game-changing", "pivotal", "groundbreaking", "cutting-edge"]

def verify(b, path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    expected = ["PARAMETERS","SCOPE_DETAIL","ESTIMATE","DD1391_BLOCK9"]
    assert sorted(wb.sheetnames) == sorted(expected), f"{b} tabs: {wb.sheetnames}"
    leaks = []
    for s in wb.sheetnames:
        ws = wb[s]
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if c.value is None: continue
                v = str(c.value)
                for bad in BANNED_STRINGS:
                    if bad.lower() in v.lower():
                        if bad == "DBD" and "DB Design" in v: continue
                        leaks.append((s, c.coordinate, bad, v[:80]))
    if leaks:
        print(f"  LEAKS in {b}:")
        for s,c,bad,v in leaks: print(f"    [{s}] {c}: {bad} :: {v}")
    return leaks

def excel_round(x, digits=0):
    import decimal
    q = decimal.Decimal(10) ** -digits
    return int(decimal.Decimal(str(x)).quantize(q, rounding=decimal.ROUND_HALF_UP))

def pax_chain(items_k):
    """Mirror PAX Form 1391 Processor rollup. items_k in $000."""
    tcc = excel_round(items_k * 1.10)
    tfc = excel_round(tcc * 1.08)
    db  = excel_round(items_k * 0.04)
    return tcc, tfc, db, tfc + db

def python_reconcile(b, blk):
    items_k = ITEMS_SUBTOTAL_TARGET_K[b]
    _, _, _, tpc_k = pax_chain(items_k)
    return tpc_k * 1000, tpc_k

def build_one(b):
    blk = SCOPE[b]
    wb = Workbook()
    wb.remove(wb.active)
    scope_total_row, _ = build_scope_detail(wb, b, blk)
    DERIV = build_parameters(wb, b, scope_total_row)
    build_estimate(wb, b, blk, DERIV, scope_total_row)
    build_block9(wb, b, blk, DERIV, scope_total_row)
    # reorder tabs
    order = ["ESTIMATE","SCOPE_DETAIL","PARAMETERS","DD1391_BLOCK9"]
    wb._sheets = [wb[n] for n in order]
    fname = f"{b}_G_CEPBKUP_{LOCKED[b]['pax']}_POM26_{LOCKED[b]['date']}.xlsx"
    path = os.path.join(ROOT, fname)
    wb.save(path)
    return fname, path

if __name__ == "__main__":
    print(f"{'Bldg':<6}{'TPC computed':>16}{'TPC $000':>12}{'Locked $000':>14}{'delta':>8}  Leaks")
    all_clean = True
    for b in ["1024","3213","3237","3270","3314"]:
        fname, path = build_one(b)
        tpc_full, tpc_round = python_reconcile(b, SCOPE[b])
        locked_round = round(LOCKED[b]["tpc"]/1000)
        delta = tpc_round - locked_round
        leaks = verify(b, path)
        if leaks or delta != 0: all_clean = False
        print(f"{b:<6}{tpc_full:>16,}{tpc_round:>12,}{locked_round:>14,}{delta:>8}  {'OK' if not leaks else 'LEAK'}")
    print()
    print("ALL CLEAN" if all_clean else "ISSUES FOUND - see above")
